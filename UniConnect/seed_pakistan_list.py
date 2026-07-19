"""
Seed Pakistan_Universities_List.xlsx into UniConnect database.
Deduplicates by slug — skips universities that already exist.
"""

import openpyxl
import re
import os
import hashlib
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL")
XLSX_PATH = os.path.join(os.path.dirname(__file__), "Pakistan_Universities_List.xlsx")
XLSX_PATH2 = os.path.join(os.path.dirname(__file__), "University_List_with_links.xlsx")


def make_slug(name):
    name = re.sub(r'\[[\d]+\]', '', name)
    slug = name.lower().strip()
    slug = re.sub(r'[,&]', '', slug)
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'\s+', '-', slug)
    slug = re.sub(r'-+', '-', slug)
    slug = slug.strip('-')
    return slug


def make_cuid(name):
    h = hashlib.md5(name.encode()).hexdigest()[:20]
    return f"clx{h}"


def normalize_province(p):
    if not p:
        return None
    p = p.strip()
    mapping = {
        "islamabad capital territory": "Islamabad",
        "kpk": "Khyber Pakhtunkhwa",
        "khyber pakhtunkhwa": "Khyber Pakhtunkhwa",
        "punjab": "Punjab",
        "sindh": "Sindh",
        "balochistan": "Balochistan",
        "azad jammu & kashmir": "Azad Kashmir",
        "azad kashmir": "Azad Kashmir",
        "gilgit-baltistan": "Gilgit-Baltistan",
        "gilgit baltistan": "Gilgit-Baltistan",
    }
    return mapping.get(p.lower(), p)


def normalize_type(t):
    if not t:
        return "PUBLIC"
    t = t.strip().upper()
    if "PRIVATE" in t:
        return "PRIVATE"
    if "MILITARY" in t:
        return "MILITARY"
    return "PUBLIC"


def read_existing_slugs(cursor):
    cursor.execute('SELECT slug FROM "University"')
    return {row[0] for row in cursor.fetchall()}


def read_pakistan_list():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    entries = []
    seen = set()

    for row_num in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[row_num]]
        name_raw = str(row[1]).strip() if row[1] else None
        if not name_raw or name_raw == "None":
            continue

        name = name_raw
        sector = normalize_type(str(row[2])) if row[2] else "PUBLIC"
        city = str(row[3]).strip() if row[3] else ""
        province = normalize_province(str(row[4]) if row[4] else None)
        website = str(row[5]).strip() if row[5] else ""
        admission = str(row[6]).strip() if row[6] else ""

        if not province:
            continue

        slug = make_slug(name)
        dedup_key = slug

        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        if website and not website.startswith("http"):
            website = f"https://{website}"
        if admission and not admission.startswith("http"):
            admission = None

        entries.append({
            "id": make_cuid(name),
            "name": name,
            "slug": slug,
            "province": province,
            "city": city if city else "Unknown",
            "type": sector,
            "websiteUrl": website if website and website != "None" else None,
            "admissionUrl": admission if admission and admission != "None" else None,
        })

    return entries


def main():
    print(f"Connecting to database...")
    conn = psycopg2.connect(DATABASE_URL)
    cursor = conn.cursor()

    existing_slugs = read_existing_slugs(cursor)
    print(f"Existing universities in DB: {len(existing_slugs)}")

    entries = read_pakistan_list()
    print(f"Total entries in Pakistan_Universities_List.xlsx: {len(entries)}")

    new_entries = [e for e in entries if e["slug"] not in existing_slugs]
    print(f"New entries to insert: {len(new_entries)}")

    if not new_entries:
        print("No new universities to add.")
        cursor.close()
        conn.close()
        return

    insert_sql = """
        INSERT INTO "University"
            ("id", "name", "slug", "province", "city", "type", "websiteUrl", "admissionUrl",
             "logoUrl", "ranking", "isActive", "isFeatured", "createdAt", "updatedAt")
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NULL, NULL, TRUE, FALSE, NOW(), NOW())
        ON CONFLICT (slug) DO NOTHING;
    """

    inserted = 0
    for e in new_entries:
        try:
            cursor.execute(insert_sql, (
                e["id"], e["name"], e["slug"], e["province"], e["city"],
                e["type"], e["websiteUrl"], e["admissionUrl"]
            ))
            if cursor.rowcount > 0:
                inserted += 1
        except Exception as ex:
            print(f"  Error inserting {e['name']}: {ex}")

    conn.commit()
    print(f"Inserted {inserted} new universities")

    # Count by province
    cursor.execute('''
        SELECT province, COUNT(*) FROM "University"
        GROUP BY province ORDER BY province
    ''')
    print("\nUniversity count by province:")
    for row in cursor.fetchall():
        print(f"  {row[0]}: {row[1]}")

    cursor.close()
    conn.close()
    print("\nDone!")


if __name__ == "__main__":
    main()
