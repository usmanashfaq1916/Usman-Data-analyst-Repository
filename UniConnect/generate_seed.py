import openpyxl
import re
import os
import hashlib

def make_cuid(name):
    """Generate a deterministic cuid-like ID from university name."""
    h = hashlib.md5(name.encode()).hexdigest()[:20]
    return f"clx{h}"

def make_slug(name):
    """Generate a URL-safe slug from university name."""
    # Remove footnote references like [5], [9], etc.
    name = re.sub(r'\[[\d]+\]', '', name)
    # Remove trailing notes like ", Pakistan"
    # Lowercase
    slug = name.lower().strip()
    # Replace special characters
    slug = re.sub(r'[,&]', '', slug)
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'\s+', '-', slug)
    slug = re.sub(r'-+', '-', slug)
    slug = slug.strip('-')
    return slug

def clean_name(name):
    """Remove footnote references from university name."""
    name = re.sub(r'\[[\d]+\]', '', name)
    name = name.strip()
    return name

def clean_year(val):
    """Extract numeric year from potentially noisy values like '1970[5]' or '1986[8]'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    m = re.match(r'(\d{4})', s)
    if m:
        return int(m.group(1))
    return None

def clean_url(val):
    """Clean URL values, return None for non-URLs."""
    if val is None:
        return None
    s = str(val).strip()
    if s.startswith('Verify') or s.startswith('N/A') or s.startswith('Bill passed') or s.startswith('Under IBA'):
        return None
    # Extract actual URL if present
    m = re.search(r'(https?://[^\s,)]+)', s)
    if m:
        return m.group(1)
    # If it looks like a domain without protocol
    if re.match(r'[\w.-]+\.(edu|com|org|pk|net)', s):
        return f'https://{s}'
    return None

def clean_campuses(val):
    """Clean campuses field, remove notes."""
    if val is None:
        return None
    s = str(val).strip()
    # Remove parenthetical notes
    s = re.sub(r'\([^)]*\)', '', s)
    s = s.strip(', ')
    if not s or s.lower() in ('none', ''):
        return None
    return s

def infer_city_from_name(name, province):
    """Try to infer city from university name."""
    name_lower = name.lower()
    city_map = {
        'peshawar': 'Peshawar', 'lahore': 'Lahore', 'karachi': 'Karachi',
        'islamabad': 'Islamabad', 'rawalpindi': 'Rawalpindi', 'multan': 'Multan',
        'faisalabad': 'Faisalabad', 'quetta': 'Quetta', 'hyderabad': 'Hyderabad',
        'sialkot': 'Sialkot', 'gujrat': 'Gujrat', 'sargodha': 'Sargodha',
        'bhawalpur': 'Bahawalpur', 'bahawalpur': 'Bahawalpur', 'taxila': 'Taxila',
        'abbottabad': 'Abbottabad', 'bannu': 'Bannu', 'kohat': 'Kohat',
        'mardan': 'Mardan', 'swat': 'Swat', 'swabi': 'Swabi',
        'charsadda': 'Charsadda', 'nowshera': 'Nowshera', 'haripur': 'Haripur',
        'mansehra': 'Mansehra', 'gilgit': 'Gilgit', 'skardu': 'Skardu',
        'mirpur': 'Mirpur', 'muzaffarabad': 'Muzaffarabad', 'bagh': 'Bagh',
        'kotli': 'Kotli', 'rawalakot': 'Rawalakot', 'tandojam': 'Tandojam',
        'jamshoro': 'Jamshoro', 'sukkur': 'Sukkur', 'larkana': 'Larkana',
        'khairpur': 'Khairpur', 'dadu': 'Dadu', 'nawabshah': 'Nawabshah',
        'benazirabad': 'Nawabshah', 'nankana': 'Nankana Sahib',
        'chakwal': 'Chakwal', 'mianwali': 'Mianwali', 'sahiwal': 'Sahiwal',
        'okara': 'Okara', 'jhang': 'Jhang', 'dg khan': 'Dera Ghazi Khan',
        'derah': 'Dera Ghazi Khan', 'rahyam': 'Rahim Yar Khan',
        'rahim yar khan': 'Rahim Yar Khan', 'wah': 'Wah', 'murree': 'Murree',
        'khanewal': 'Khanewal', 'layyah': 'Layyah', 'bhakkar': 'Bhakkar',
        'narowal': 'Narowal', 'gujranwala': 'Gujranwala',
        'turbat': 'Turbat', 'gwadar': 'Gwadar', 'khuzdar': 'Khuzdar',
        'loralai': 'Loralai', 'sibi': 'Sibi', 'panjgur': 'Panjgur',
        'lasbela': 'Lasbela', 'nasirabad': 'Nasirabad',
        'nagarparkar': 'Nagarparkar', 'tharparkar': 'Tharparkar',
        'mirpurkhas': 'Mirpurkhas', 'kamalia': 'Kamalia',
        'bhimber': 'Bhimber', 'hunza': 'Hunza',
    }
    # Check if any city name appears in the university name
    for key, city in city_map.items():
        if key in name_lower:
            return city
    # Province fallback cities
    province_capital = {
        'Islamabad': 'Islamabad', 'Balochistan': 'Quetta',
        'Khyber Pakhtunkhwa': 'Peshawar', 'Punjab': 'Lahore',
        'Sindh': 'Karachi', 'Azad Kashmir': 'Mirpur',
        'Gilgit-Baltistan': 'Gilgit',
    }
    return province_capital.get(province, 'Unknown')

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(script_dir, 'University_List_with_links.xlsx')
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb.active

    # Province section header rows and their names
    province_sections = {
        1: 'Islamabad',
        31: 'Balochistan',
        58: 'Khyber Pakhtunkhwa',
        107: 'Punjab',
        211: 'Sindh',
        296: 'Azad Kashmir',
        310: 'Gilgit-Baltistan',
    }

    # Rows to skip (metadata, notes, headers, supplementary data)
    skip_rows = set()
    # Add empty rows
    for row_num in range(1, sheet.max_row + 1):
        row = [cell.value for cell in sheet[row_num]]
        if all(v is None for v in row):
            skip_rows.add(row_num)

    # Add province header rows, excerpt rows, column header rows
    skip_rows.update([1, 2, 3, 4, 31, 32, 33, 34, 58, 59, 60, 61,
                      107, 108, 109, 110, 211, 212, 213, 214,
                      296, 297, 298, 299, 310, 311, 312])

    # Skip known non-university rows
    skip_rows.update([41, 44, 46, 48, 56, 269, 291, 308])

    universities = []
    seen_names = set()
    current_province = 'Islamabad'

    for row_num in range(1, sheet.max_row + 1):
        # Detect province transitions FIRST (before skip_rows)
        if row_num in province_sections:
            current_province = province_sections[row_num]
            continue

        if row_num in skip_rows:
            continue

        row = [cell.value for cell in sheet[row_num]]

        # Skip rows with no university name
        if not row[0] or not isinstance(row[0], str):
            continue

        name_raw = str(row[0]).strip()

        # Skip if it looks like a header or note
        if name_raw in ('University', 'Location', 'Established', 'Specialization', 'Type'):
            continue
        if 'excerpt from' in name_raw.lower():
            continue
        if name_raw.startswith('* '):
            continue
        if 'purposed' in name_raw.lower() or 'has two campuses' in name_raw.lower():
            continue

        # Determine column layout based on province
        # Islamabad: [Name, Established, Campuses, Specialization, Type, _, Website, Admission]
        # Balochistan onwards: [Name, Location, Established, Campuses, Specialization, Type, Website, Admission]

        if current_province == 'Islamabad':
            established = row[1]
            campuses = row[2]
            specialization = row[3]
            uni_type = row[4]
            website = row[6]
            admission = row[7]
        else:
            # Has Location column
            established = row[2]
            campuses = row[3]
            specialization = row[4]
            uni_type = row[5]
            website = row[6]
            admission = row[7]

        # Clean values
        name = clean_name(name_raw)
        year = clean_year(established)
        uni_type_str = str(uni_type).strip() if uni_type else None
        if uni_type_str and uni_type_str.lower() in ('none', ''):
            uni_type_str = None

        # Determine city
        if current_province == 'Islamabad':
            city = 'Islamabad'
        else:
            city = row[1]  # Location column
            if not city or not isinstance(city, str):
                city = infer_city_from_name(name, current_province)
            else:
                city = city.strip()

        # Skip clearly non-university entries
        if 'HEC-NOC SUSPENDED' in name_raw.upper():
            continue

        # Skip proposed/purposed entries with no real data
        established_str = str(established).lower() if established else ''
        if 'purposed' in established_str:
            continue

        website_url = clean_url(website)
        admission_url = clean_url(admission)

        # Skip entries that are clearly not universities (no type, no URLs, no year)
        if not uni_type_str and not website_url and not year:
            continue

        # Clean city names
        if current_province != 'Islamabad' and city:
            city = city.split('@')[0].split('[')[0].strip()

        uni_data = {
            'name': name,
            'slug': make_slug(name),
            'province': current_province,
            'city': city if city else infer_city_from_name(name, current_province),
            'type': uni_type_str if uni_type_str else 'General',
            'websiteUrl': website_url,
            'admissionUrl': admission_url,
        }

        # Deduplicate by name (keep first occurrence)
        name_key = name.lower().strip()
        if name_key in seen_names:
            print(f'  Skipping duplicate: {name} ({current_province})')
            continue
        seen_names.add(name_key)

        universities.append(uni_data)

    # Generate SQL
    lines = []
    lines.append('-- UniConnect Seed Data: Universities')
    lines.append(f'-- Generated from University_List_with_links.xlsx')
    lines.append(f'-- Total universities: {len(universities)}')
    lines.append('')
    lines.append('BEGIN;')
    lines.append('')

    for uni in universities:
        id_val = make_cuid(uni['name'])
        name_val = uni['name'].replace("'", "''")
        slug_val = uni['slug']
        province_val = uni['province']
        city_val = uni['city'].replace("'", "''") if uni['city'] else 'Unknown'
        type_val = uni['type'] if uni['type'] else 'General'
        website_val = f"'{uni['websiteUrl']}'" if uni['websiteUrl'] else 'NULL'
        admission_val = f"'{uni['admissionUrl']}'" if uni['admissionUrl'] else 'NULL'

        sql = (
            f"INSERT INTO \"University\" (\"id\", \"name\", \"slug\", \"province\", \"city\", "
            f"\"type\", \"websiteUrl\", \"admissionUrl\", \"logoUrl\", \"ranking\", \"createdAt\") "
            f"VALUES ('{id_val}', '{name_val}', '{slug_val}', '{province_val}', '{city_val}', "
            f"'{type_val}', {website_val}, {admission_val}, NULL, NULL, NOW());"
        )
        lines.append(sql)

    lines.append('')
    lines.append('COMMIT;')
    lines.append('')
    lines.append(f'-- Note: Program and Admission tables need separate seeding.')
    lines.append(f'-- University count by province:')

    # Count by province
    from collections import Counter
    province_counts = Counter(u['province'] for u in universities)
    for prov, count in sorted(province_counts.items()):
        lines.append(f'--   {prov}: {count}')

    output = '\n'.join(lines)

    output_path = os.path.join(script_dir, 'seed.sql')
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(output)

    print(f'Generated seed.sql with {len(universities)} universities')
    for prov, count in sorted(province_counts.items()):
        print(f'  {prov}: {count}')

if __name__ == '__main__':
    main()
